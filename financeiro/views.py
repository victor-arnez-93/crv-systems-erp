from django.shortcuts import render
from django.http import HttpResponse
from django.template.loader import render_to_string
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Image, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
import io
import os
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import datetime
from .models import MovimentacaoFinanceira


from .models import MovimentacaoFinanceira
from django.utils import timezone

def financeiro_view(request):
    movs = MovimentacaoFinanceira.objects.all()
    # Filtros
    periodo = request.GET.get('periodo', '')
    tipo    = request.GET.get('tipo', '')
    data_de = request.GET.get('data_de', '')
    data_ate= request.GET.get('data_ate', '')

    hoje = timezone.now().date()
    if periodo == 'hoje':
        movs = movs.filter(data=hoje)
    elif periodo == 'semana':
        semana = hoje - datetime.timedelta(days=7)
        movs = movs.filter(data__gte=semana, data__lte=hoje)
    elif periodo == 'mes':
        mes = hoje.replace(day=1)
        movs = movs.filter(data__gte=mes, data__lte=hoje)
    elif periodo == 'personalizado':
        if data_de:
            movs = movs.filter(data__gte=data_de)
        if data_ate:
            movs = movs.filter(data__lte=data_ate)
    if tipo in ['entrada', 'saida']:
        movs = movs.filter(tipo=tipo)
    movs = movs.order_by('-data')

    total_entradas = sum(x.valor for x in movs if x.tipo == 'entrada')
    total_saidas = sum(x.valor for x in movs if x.tipo == 'saida')
    saldo = total_entradas - total_saidas

    context = {
        'movimentacoes': movs,
        'total_entradas': total_entradas,
        'total_saidas': total_saidas,
        'saldo': saldo,
    }
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return render(request, 'financeiro/financeiro_parcial.html', context)
    context['partial'] = render_to_string('financeiro/financeiro_parcial.html', context=context, request=request)
    return render(request, 'base.html', context)

def exportar_pdf(request):
    from .models import MovimentacaoFinanceira
    from django.utils import timezone

    # Filtros iguais à view principal
    movs = MovimentacaoFinanceira.objects.all()
    periodo = request.GET.get('periodo', '')
    tipo    = request.GET.get('tipo', '')
    data_de = request.GET.get('data_de', '')
    data_ate= request.GET.get('data_ate', '')

    hoje = timezone.now().date()
    if periodo == 'hoje':
        movs = movs.filter(data=hoje)
    elif periodo == 'semana':
        semana = hoje - datetime.timedelta(days=7)
        movs = movs.filter(data__gte=semana, data__lte=hoje)
    elif periodo == 'mes':
        mes = hoje.replace(day=1)
        movs = movs.filter(data__gte=mes, data__lte=hoje)
    elif periodo == 'personalizado':
        if data_de:
            movs = movs.filter(data__gte=data_de)
        if data_ate:
            movs = movs.filter(data__lte=data_ate)
    if tipo in ['entrada', 'saida']:
        movs = movs.filter(tipo=tipo)
    movs = movs.order_by('-data')

    total_entradas = sum(x.valor for x in movs if x.tipo == 'entrada')
    total_saidas = sum(x.valor for x in movs if x.tipo == 'saida')
    saldo = total_entradas - total_saidas

    # Gráfico matplotlib com dados filtrados
    import io, os
    import matplotlib.pyplot as plt
    plt.figure(figsize=(6,2.3))
    plt.bar(['Entradas','Saídas','Saldo'], [total_entradas, total_saidas, saldo], color=['#33cc4c','#d94343','#0099ff'])
    plt.ylabel('Valor (R$)')
    plt.title('Balanço Financeiro')
    for i, v in enumerate([total_entradas, total_saidas, saldo]):
        plt.text(i, v+200, f'R$ {v:,.2f}', ha='center', fontsize=8)
    plt.tight_layout()
    buffer_graf = io.BytesIO()
    plt.savefig(buffer_graf, format='png', dpi=140, transparent=False)
    plt.close()
    buffer_graf.seek(0)

    # PDF reportlab
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Image, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()

    # Caminho absoluto da logo CRV
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    logo_path = os.path.join(BASE_DIR, 'static', 'icones', 'logo_crv.png')
    if os.path.exists(logo_path):
        elements.append(Image(logo_path, width=170))
    else:
        # fallback logo antiga se não encontrar
        logo2 = os.path.join(BASE_DIR, 'static', 'icones', 'logo_crv2.png')
        if os.path.exists(logo2):
            elements.append(Image(logo2, width=170))

    elements.append(Spacer(1, 12))
    data_geracao = datetime.date.today().strftime('%d/%m/%Y')
    elements.append(Paragraph(
        f'<para align="left"><font size=9><i>Data de geração: {data_geracao}</i></font></para>',
        styles['Normal']
    ))
    elements.append(Spacer(1, 6))
    title_style = styles['Heading1']
    title_style.alignment = 1
    elements.append(Paragraph('Relatório Financeiro', title_style))
    elements.append(Spacer(1, 8))
    graf_img = Image(buffer_graf, width=400, height=155)
    graf_img.hAlign = 'CENTER'
    elements.append(graf_img)
    elements.append(Spacer(1, 8))

    # Tabela
    data_tabela = [['#','Data','Tipo','Valor','Descrição']]
    for idx, d in enumerate(movs, start=1):
        data_tabela.append([
            idx,
            d.data.strftime('%d/%m/%Y'),
            d.get_tipo_display(),
            f"R$ {d.valor:,.2f}".replace('.',','),
            f"{d.descricao or ''} {(f'({d.produto_nome})' if d.produto_nome else '')}"
        ])
    t = Table(data_tabela, colWidths=[20, 75, 55, 70, 130])
    t.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,0), colors.HexColor('#2973b5')),
        ('TEXTCOLOR',(0,0),(-1,0), colors.white),
        ('GRID',(0,0),(-1,-1), 0.4, colors.black),
        ('ALIGN',(0,0),(-1,-1),'CENTER'),
        ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
        ('FONTSIZE',(0,0),(-1,-1), 9),
    ]))
    elements.append(t)
    elements.append(Spacer(1, 8))

    # Totais
    elements.append(Paragraph(
        f"<b>Total de Entradas:</b> R$ {total_entradas:,.2f}<br/>"
        f"<b>Total de Saídas:</b> R$ {total_saidas:,.2f}<br/>"
        f"<b>Saldo Final:</b> R$ {saldo:,.2f}",
        styles['Normal']
    ))

    doc.build(elements)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="relatorio.pdf"'
    return response

def exportar_excel(request):
    from .models import MovimentacaoFinanceira
    from django.utils import timezone

    movs = MovimentacaoFinanceira.objects.all()
    periodo = request.GET.get('periodo', '')
    tipo    = request.GET.get('tipo', '')
    data_de = request.GET.get('data_de', '')
    data_ate= request.GET.get('data_ate', '')

    hoje = timezone.now().date()
    if periodo == 'hoje':
        movs = movs.filter(data=hoje)
    elif periodo == 'semana':
        semana = hoje - datetime.timedelta(days=7)
        movs = movs.filter(data__gte=semana, data__lte=hoje)
    elif periodo == 'mes':
        mes = hoje.replace(day=1)
        movs = movs.filter(data__gte=mes, data__lte=hoje)
    elif periodo == 'personalizado':
        if data_de:
            movs = movs.filter(data__gte=data_de)
        if data_ate:
            movs = movs.filter(data__lte=data_ate)
    if tipo in ['entrada', 'saida']:
        movs = movs.filter(tipo=tipo)
    movs = movs.order_by('-data')

    total_entradas = sum(x.valor for x in movs if x.tipo == 'entrada')
    total_saidas = sum(x.valor for x in movs if x.tipo == 'saida')
    saldo = total_entradas - total_saidas

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório"
    ws.append(["#", "Data", "Tipo", "Valor", "Descrição"])
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2973b5")
        cell.alignment = Alignment(horizontal='center')

    for idx, d in enumerate(movs, 1):
        ws.append([
            idx,
            d.data.strftime('%d/%m/%Y'),
            d.get_tipo_display(),
            float(d.valor),
            f"{d.descricao or ''} {(f'({d.produto_nome})' if d.produto_nome else '')}"
        ])

    ws.append([])
    ws.append(['', '', '', 'Resumo Financeiro', ''])
    ws.append(['', '', '', 'Total de Entradas', f"R$ {total_entradas:,.2f}".replace('.',',')])
    ws.append(['', '', '', 'Total de Saídas', f"R$ {total_saidas:,.2f}".replace('.',',')])
    ws.append(['', '', '', 'Saldo Final', f"R$ {saldo:,.2f}".replace('.',',')])

    for row in ws.iter_rows(min_row=ws.max_row-2, max_row=ws.max_row):
        for cell in row:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='right')

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="relatorio.xlsx"'
    return response
